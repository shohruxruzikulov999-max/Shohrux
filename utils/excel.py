"""
utils/excel.py — buyurtmalarni Excel ga export qilish
"""
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter


def orders_to_excel(orders: list) -> bytes:
    """
    orders — Order modellar ro'yxati
    qaytaradi: bytes (Excel fayl)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Buyurtmalar"

    # ── Ranglar ──────────────────────────────────────────────────────────
    HEADER_FILL  = PatternFill("solid", fgColor="4F7EF8")
    ALT_FILL     = PatternFill("solid", fgColor="EEF2FF")
    GREEN_FILL   = PatternFill("solid", fgColor="D1FAE5")
    YELLOW_FILL  = PatternFill("solid", fgColor="FEF3C7")
    RED_FILL     = PatternFill("solid", fgColor="FEE2E2")
    TOTAL_FILL   = PatternFill("solid", fgColor="1D4ED8")

    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Sarlavha ─────────────────────────────────────────────────────────
    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value = f"📦 BUYURTMALAR RO'YXATI — {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill = HEADER_FILL
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # ── Ustun sarlavhalari ────────────────────────────────────────────────
    headers = [
        "№", "Buyurtma #", "Foydalanuvchi ID",
        "Mahsulot", "Miqdor", "Narx (so'm)",
        "Mijoz ma'lumotlari", "Status", "Sana"
    ]
    widths = [5, 12, 18, 22, 8, 14, 40, 14, 18]

    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[2].height = 24

    # ── Ma'lumotlar ───────────────────────────────────────────────────────
    status_map = {
        "pending":   ("⏳ Kutmoqda",   YELLOW_FILL),
        "confirmed": ("✅ Tasdiqlandi", GREEN_FILL),
        "cancelled": ("❌ Bekor",       RED_FILL),
    }

    for i, o in enumerate(orders, 1):
        row = i + 2
        fill = ALT_FILL if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")

        status_txt, status_fill = status_map.get(o.status, (o.status, fill))

        values = [
            i,
            f"#{o.id}",
            o.user_id,
            o.product,
            o.quantity,
            int(o.price),
            o.note or "—",
            status_txt,
            o.created_at.strftime("%d.%m.%Y %H:%M") if o.created_at else "—",
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = border
            cell.alignment = Alignment(
                vertical="center",
                horizontal="center" if col != 7 else "left",
                wrap_text=True
            )
            cell.font = Font(size=10)

            if col == 8:
                cell.fill = status_fill
                cell.font = Font(size=10, bold=True)
            elif col == 6:
                cell.font = Font(size=10, bold=True, color="1D4ED8")
                cell.fill = fill
            else:
                cell.fill = fill

        ws.row_dimensions[row].height = 20

    # ── Jami qator ───────────────────────────────────────────────────────
    total_row = len(orders) + 3
    confirmed_orders = [o for o in orders if o.status == "confirmed"]
    total_sum = sum(o.price for o in confirmed_orders)
    total_count = len(orders)

    ws.merge_cells(f"A{total_row}:E{total_row}")
    sum_label = ws.cell(row=total_row, column=1,
                        value=f"JAMI: {total_count} ta buyurtma | Tasdiqlangan: {len(confirmed_orders)} ta")
    sum_label.font = Font(bold=True, size=11, color="FFFFFF")
    sum_label.fill = TOTAL_FILL
    sum_label.alignment = Alignment(horizontal="center", vertical="center")
    sum_label.border = border

    ws.merge_cells(f"F{total_row}:I{total_row}")
    sum_val = ws.cell(row=total_row, column=6,
                      value=f"{int(total_sum):,} so'm (tasdiqlangan)")
    sum_val.font = Font(bold=True, size=11, color="FFFFFF")
    sum_val.fill = TOTAL_FILL
    sum_val.alignment = Alignment(horizontal="center", vertical="center")
    sum_val.border = border
    ws.row_dimensions[total_row].height = 26

    # ── Freeze ───────────────────────────────────────────────────────────
    ws.freeze_panes = "A3"

    # ── Bytes ga aylantirish ─────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()